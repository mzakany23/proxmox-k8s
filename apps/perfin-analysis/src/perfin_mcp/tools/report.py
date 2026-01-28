"""Report generation tool for MCP - creates Excel financial reports."""

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side

from .spending import get_spending_analysis
from .fixed_costs import get_fixed_costs


def _get_reports_dir() -> Path:
    """Get the reports directory."""
    mcp_server_dir = Path(__file__).parent.parent.parent.parent
    project_root = mcp_server_dir.parent
    reports_dir = project_root / "reports"
    reports_dir.mkdir(exist_ok=True)
    return reports_dir


def _create_styles(wb: Workbook) -> dict:
    """Create and register named styles for the workbook."""
    styles = {}

    # Header style
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF", size=11)
    header_style.fill = PatternFill("solid", fgColor="2F5496")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(bottom=Side(style="medium", color="000000"))
    wb.add_named_style(header_style)
    styles["header"] = header_style

    # Section header style
    section_style = NamedStyle(name="section")
    section_style.font = Font(bold=True, size=12, color="2F5496")
    section_style.fill = PatternFill("solid", fgColor="D6DCE5")
    section_style.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(section_style)
    styles["section"] = section_style

    # Currency style
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = '"$"#,##0.00'
    currency_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(currency_style)
    styles["currency"] = currency_style

    # Percentage style
    percent_style = NamedStyle(name="percent")
    percent_style.number_format = "0.0%"
    percent_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(percent_style)
    styles["percent"] = percent_style

    # Total row style
    total_style = NamedStyle(name="total")
    total_style.font = Font(bold=True, size=11)
    total_style.fill = PatternFill("solid", fgColor="E2EFDA")
    total_style.number_format = '"$"#,##0.00'
    total_style.border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )
    wb.add_named_style(total_style)
    styles["total"] = total_style

    # Negative currency style
    negative_style = NamedStyle(name="negative")
    negative_style.font = Font(color="CC0000")
    negative_style.number_format = '"$"#,##0.00'
    negative_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(negative_style)
    styles["negative"] = negative_style

    return styles


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            try:
                if hasattr(cell, "column_letter"):
                    if column is None:
                        column = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        if column:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width


def _create_summary_sheet(wb: Workbook, spending: dict, fixed: dict) -> None:
    """Create the Key Metrics Summary sheet."""
    ws = wb.active
    ws.title = "Key Metrics"

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value=f"Financial Summary (Generated {datetime.now().strftime('%Y-%m-%d')})")
    title_cell.font = Font(bold=True, size=16, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    # Subtitle with data range
    ws.merge_cells("A2:C2")
    subtitle = ws.cell(row=2, column=1, value=f"Based on {spending.get('months_analyzed', 6)}-month averages")
    subtitle.font = Font(italic=True, size=10, color="666666")
    subtitle.alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Metric", "Value", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.style = "header"

    # Extract data
    income = spending["totals"]["income"]
    fixed_total = spending["totals"]["fixed"]
    variable = spending["totals"]["variable"]
    surplus = spending["totals"]["surplus"]
    savings_rate = spending["savings_rate"] / 100

    # Post-June projection (single income ~$12k)
    projected_income = 12000
    projected_surplus = projected_income - fixed_total - variable

    data = [
        ("MONTHLY CASH FLOW (Current)", "", ""),
        ("Monthly Take-Home", income, f"Avg from {spending['months_analyzed']} months"),
        ("Fixed Costs", fixed_total, f"{fixed['months_analyzed']}-month average"),
        ("Variable Spending", variable, "By category below"),
        ("Monthly Surplus", surplus, f"{savings_rate:.0%} savings rate"),
        ("", "", ""),
        ("POST-JUNE 2026 PROJECTION", "", ""),
        ("Income (Single)", projected_income, "Musical Arts ends June"),
        ("Expenses (unchanged)", fixed_total + variable, ""),
        ("Monthly Shortfall", projected_surplus, "Need spending cuts" if projected_surplus < 0 else "Sustainable"),
        ("", "", ""),
        ("KEY RATIOS", "", ""),
        ("Savings Rate", savings_rate, "Of take-home income"),
        ("Housing Ratio", fixed_total / income if income else 0, "Fixed costs / income"),
    ]

    row = 5
    for metric, value, notes in data:
        ws.cell(row=row, column=1, value=metric)
        if value != "":
            cell = ws.cell(row=row, column=2, value=value)
            if isinstance(value, (int, float)):
                if metric == "Savings Rate" or metric == "Housing Ratio":
                    cell.style = "percent"
                elif value < 0:
                    cell.style = "negative"
                elif "TOTAL" in metric or "Surplus" in metric or "Shortfall" in metric:
                    cell.style = "total"
                else:
                    cell.style = "currency"
        ws.cell(row=row, column=3, value=notes)

        # Section headers
        if metric in ("MONTHLY CASH FLOW (Current)", "POST-JUNE 2026 PROJECTION", "KEY RATIOS"):
            ws.cell(row=row, column=1).style = "section"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

        row += 1

    _auto_width(ws)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30


def _create_spending_sheet(wb: Workbook, spending: dict) -> None:
    """Create the Variable Spending sheet."""
    ws = wb.create_sheet("Variable Spending")

    # Headers
    headers = ["Category", "Monthly Avg", "% of Total", "6-Month Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "header"

    row = 2
    for cat in spending["categories"]:
        ws.cell(row=row, column=1, value=cat["name"])
        cell = ws.cell(row=row, column=2, value=cat["monthly_avg"])
        cell.style = "currency"
        cell = ws.cell(row=row, column=3, value=cat["percent"] / 100)
        cell.style = "percent"
        cell = ws.cell(row=row, column=4, value=cat["total"])
        cell.style = "currency"
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL")
    cell = ws.cell(row=row, column=2, value=spending["totals"]["variable"])
    cell.style = "total"
    ws.cell(row=row, column=3, value=1.0).style = "percent"

    _auto_width(ws)


def _create_fixed_costs_sheet(wb: Workbook, fixed: dict) -> None:
    """Create the Fixed Costs sheet."""
    ws = wb.create_sheet("Fixed Costs")

    # Headers
    headers = ["Category", "Monthly Amount", "Annual", "Primary Merchant"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "header"

    row = 2
    for cost in fixed["costs"]:
        ws.cell(row=row, column=1, value=cost["category"])
        cell = ws.cell(row=row, column=2, value=cost["amount"])
        cell.style = "currency"
        cell = ws.cell(row=row, column=3, value=cost["amount"] * 12)
        cell.style = "currency"
        ws.cell(row=row, column=4, value=cost["merchant"])
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL")
    cell = ws.cell(row=row, column=2, value=fixed["total_monthly"])
    cell.style = "total"
    cell = ws.cell(row=row, column=3, value=fixed["total_annual"])
    cell.style = "total"

    _auto_width(ws)


def _create_budget_scenarios_sheet(wb: Workbook, spending: dict, fixed: dict) -> None:
    """Create the Budget Scenarios sheet with what-if analysis."""
    ws = wb.create_sheet("Budget Scenarios")

    income = spending["totals"]["income"]
    fixed_total = spending["totals"]["fixed"]
    variable = spending["totals"]["variable"]

    # Headers
    headers = ["Scenario", "Income", "Fixed", "Variable", "Surplus", "Savings Rate"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = "header"

    scenarios = [
        ("Current (Dual Income)", income, fixed_total, variable),
        ("Post-June (Single)", 12000, fixed_total, variable),
        ("10% Variable Cut", 12000, fixed_total, variable * 0.9),
        ("20% Variable Cut", 12000, fixed_total, variable * 0.8),
        ("30% Variable Cut", 12000, fixed_total, variable * 0.7),
        ("40% Variable Cut", 12000, fixed_total, variable * 0.6),
        ("50% Variable Cut", 12000, fixed_total, variable * 0.5),
    ]

    row = 2
    for name, inc, fix, var in scenarios:
        surplus = inc - fix - var
        rate = surplus / inc if inc else 0

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=inc).style = "currency"
        ws.cell(row=row, column=3, value=fix).style = "currency"
        ws.cell(row=row, column=4, value=var).style = "currency"

        cell = ws.cell(row=row, column=5, value=surplus)
        cell.style = "total" if surplus >= 0 else "negative"

        ws.cell(row=row, column=6, value=rate).style = "percent"
        row += 1

    _auto_width(ws)


def generate_report(
    output_filename: str | None = None,
    include_scenarios: bool = True
) -> dict[str, Any]:
    """
    Generate a formatted Excel financial report.

    Args:
        output_filename: Optional custom filename (default: financial-report-YYYYMMDD.xlsx)
        include_scenarios: Whether to include budget scenario analysis sheet

    Returns:
        Dictionary with file_path and sheets included
    """
    # Get live data
    spending = get_spending_analysis(months=6)
    if "error" in spending:
        return spending

    fixed = get_fixed_costs()
    if "error" in fixed:
        return fixed

    # Create workbook
    wb = Workbook()
    _create_styles(wb)

    # Create sheets
    _create_summary_sheet(wb, spending, fixed)
    _create_spending_sheet(wb, spending)
    _create_fixed_costs_sheet(wb, fixed)

    if include_scenarios:
        _create_budget_scenarios_sheet(wb, spending, fixed)

    # Save
    reports_dir = _get_reports_dir()
    if output_filename:
        filename = output_filename if output_filename.endswith(".xlsx") else f"{output_filename}.xlsx"
    else:
        filename = f"financial-report-{datetime.now().strftime('%Y%m%d')}.xlsx"

    output_path = reports_dir / filename
    wb.save(output_path)

    return {
        "file_path": str(output_path),
        "sheets": wb.sheetnames,
        "generated_at": datetime.now().isoformat(),
        "data_summary": {
            "months_analyzed": spending["months_analyzed"],
            "categories": len(spending["categories"]),
            "total_variable": spending["totals"]["variable"],
            "total_fixed": fixed["total_monthly"],
        },
    }
