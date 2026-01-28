"""Category analysis sheet."""

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseSheet


class CategoriesSheet(BaseSheet):
    """Hierarchical category analysis."""

    sheet_id = "categories"
    sheet_name = "Categories"

    def can_generate(self) -> tuple[bool, str]:
        if not self.data.has_transactions:
            return False, "No transaction data available"
        return True, ""

    def generate(self, wb: Workbook) -> Worksheet:
        ws = self._create_sheet(wb)

        df = self.data.transactions.copy()

        if df.empty:
            ws.cell(row=1, column=1, value="No transactions found")
            return ws

        # Calculate number of months for averages
        months_count = df["date"].dt.to_period("M").nunique()
        months_count = max(months_count, 1)

        # Separate expenses and income
        expenses = df[df["amount"] < 0].copy()
        expenses["abs_amount"] = expenses["amount"].abs()

        income = df[df["amount"] > 0].copy()

        # Write header
        row = 1
        headers = ["Group", "Category", "Total", "Monthly Avg", "% of Total", "Count"]
        for i, header in enumerate(headers, start=1):
            ws.cell(row=row, column=i, value=header).style = "header"
        row += 1

        # Expenses section
        self.formatter.format_section_header(ws, row, "EXPENSES", len(headers))
        row += 1

        if not expenses.empty:
            expense_total = expenses["abs_amount"].sum()
            row = self._write_category_section(ws, expenses, row, expense_total, months_count, "abs_amount", len(headers))

        # Add a blank row
        row += 1

        # Income section
        self.formatter.format_section_header(ws, row, "INCOME", len(headers))
        row += 1

        if not income.empty:
            income_total = income["amount"].sum()
            row = self._write_category_section(ws, income, row, income_total, months_count, "amount", len(headers))

        # Freeze panes
        self.formatter.freeze_panes(ws, row=2, col=1)

        # Auto-width columns
        self._auto_width(ws)

        return ws

    def _write_category_section(
        self, ws: Worksheet, df: pd.DataFrame, row: int, total: float, months_count: int, amount_col: str, num_cols: int
    ) -> int:
        """Write a section of category data."""

        group_col = "category_group" if "category_group" in df.columns else None

        if group_col:
            # Hierarchical grouping
            grouped = df.groupby([group_col, "category_name"]).agg({amount_col: ["sum", "count"]}).reset_index()
            grouped.columns = [group_col, "category_name", "total", "count"]
            grouped = grouped.sort_values("total", ascending=False)

            current_group = None
            for _, cat_row in grouped.iterrows():
                group = cat_row[group_col]
                category = cat_row["category_name"]
                cat_total = cat_row["total"]
                cat_count = int(cat_row["count"])

                # Write group name if new group
                if group != current_group:
                    current_group = group
                    ws.cell(row=row, column=1, value=group if group else "Other")

                # Write category data
                ws.cell(row=row, column=2, value=category)

                cell = ws.cell(row=row, column=3, value=cat_total)
                cell.style = "currency"

                cell = ws.cell(row=row, column=4, value=cat_total / months_count)
                cell.style = "currency"

                cell = ws.cell(row=row, column=5, value=cat_total / total if total > 0 else 0)
                cell.style = "percent"

                ws.cell(row=row, column=6, value=cat_count)
                row += 1

        else:
            # Flat grouping by category only
            grouped = df.groupby("category_name").agg({amount_col: ["sum", "count"]}).reset_index()
            grouped.columns = ["category_name", "total", "count"]
            grouped = grouped.sort_values("total", ascending=False)

            for _, cat_row in grouped.iterrows():
                ws.cell(row=row, column=2, value=cat_row["category_name"])

                cell = ws.cell(row=row, column=3, value=cat_row["total"])
                cell.style = "currency"

                cell = ws.cell(row=row, column=4, value=cat_row["total"] / months_count)
                cell.style = "currency"

                cell = ws.cell(row=row, column=5, value=cat_row["total"] / total if total > 0 else 0)
                cell.style = "percent"

                ws.cell(row=row, column=6, value=int(cat_row["count"]))
                row += 1

        # Write section total
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=1).style = "total"

        cell = ws.cell(row=row, column=3, value=total)
        cell.style = "total"

        cell = ws.cell(row=row, column=4, value=total / months_count)
        cell.style = "total"

        cell = ws.cell(row=row, column=5, value=1.0)
        cell.style = "total"

        cell = ws.cell(row=row, column=6, value=len(df))
        cell.style = "total"

        row += 1

        return row
