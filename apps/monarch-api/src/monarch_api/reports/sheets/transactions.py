"""Transaction details sheet."""

from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill

from .base import BaseSheet


class TransactionsSheet(BaseSheet):
    """Detailed transaction list with AutoFilter."""

    sheet_id = "transactions"
    sheet_name = "Transactions"

    def can_generate(self) -> tuple[bool, str]:
        if not self.data.has_transactions:
            return False, "No transaction data available"
        return True, ""

    def generate(self, wb: Workbook) -> Worksheet:
        ws = self._create_sheet(wb)

        df = self.data.transactions.copy()

        if df.empty:
            ws.cell(row=1, column=1, value="No transactions found")
            return ws

        # Sort by date descending
        df = df.sort_values("date", ascending=False)

        # Define columns to include
        columns = [
            ("date", "Date"),
            ("merchant_name", "Merchant"),
            ("category_name", "Category"),
            ("category_group", "Group"),
            ("account_name", "Account"),
            ("amount", "Amount"),
            ("tags", "Tags"),
            ("notes", "Notes"),
        ]

        # Filter to available columns
        available_cols = [(col, header) for col, header in columns if col in df.columns]

        # Write header
        row = 1
        for i, (_, header) in enumerate(available_cols, start=1):
            ws.cell(row=row, column=i, value=header).style = "header"
        row += 1

        # Alternating row colors
        alt_fill = PatternFill("solid", fgColor="F2F2F2")

        # Write data
        for idx, (_, txn) in enumerate(df.iterrows()):
            for i, (col, _) in enumerate(available_cols, start=1):
                value = txn.get(col)

                # Handle date formatting
                if col == "date" and value is not None:
                    value = value.strftime("%Y-%m-%d") if hasattr(value, "strftime") else str(value)

                cell = ws.cell(row=row, column=i, value=value)

                # Apply currency style to amount
                if col == "amount" and value is not None:
                    self.formatter.apply_currency(cell, value)

                # Alternating row colors
                if idx % 2 == 1:
                    cell.fill = alt_fill

            row += 1

        # Apply AutoFilter
        last_col = get_column_letter(len(available_cols))
        ws.auto_filter.ref = f"A1:{last_col}{row - 1}"

        # Freeze panes
        self.formatter.freeze_panes(ws, row=2, col=1)

        # Auto-width columns
        self._auto_width(ws)

        return ws
