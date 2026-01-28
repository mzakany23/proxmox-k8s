"""Summary dashboard sheet."""

from datetime import datetime

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseSheet


class SummarySheet(BaseSheet):
    """Summary dashboard with key metrics - styled like the analysis project."""

    sheet_id = "summary"
    sheet_name = "Key Metrics"

    def can_generate(self) -> tuple[bool, str]:
        if not self.data.has_transactions:
            return False, "No transaction data available"
        return True, ""

    def generate(self, wb: Workbook) -> Worksheet:
        # Create summary sheet (will be first since it's generated first)
        ws = self._create_sheet(wb)

        # Title
        start_date, end_date = self._get_date_range()
        ws.merge_cells("A1:C1")
        title_cell = ws.cell(row=1, column=1, value=f"Financial Summary - {end_date[:7]}")
        title_cell.style = "title"

        # Headers
        headers = ["Metric", "Value", "Notes"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header).style = "header"

        totals = self._calculate_totals()
        row = 4

        # Build data with sections
        data = [
            ("NET WORTH", None, None, "section"),
        ]

        # Add account balances if available
        if self.data.has_accounts:
            total_assets, total_liabilities, net_worth = self._calculate_net_worth()
            data.extend([
                ("Total Assets", total_assets, "Cash + Investments"),
                ("Total Liabilities", total_liabilities, "Credit cards, loans"),
                ("Net Worth", net_worth, "", "total"),
            ])
        else:
            data.append(("Account data not available", None, "Run monarch-export"))

        data.extend([
            (None, None, None, "spacer"),
            ("MONTHLY CASH FLOW", None, None, "section"),
            ("Total Income", totals["income"], f"Period: {start_date} to {end_date}"),
            ("Total Expenses", totals["expenses"], "Excluding transfers"),
            ("Net Surplus/Deficit", totals["net"], "", "total"),
            ("Savings Rate", totals["savings_rate"], "Net / Income", "percent"),
            (None, None, None, "spacer"),
            ("TOP SPENDING CATEGORIES", None, None, "section"),
        ])

        # Top categories
        top_categories = self._get_top_categories()
        for cat, amount, pct in top_categories:
            data.append((cat, amount, f"{pct:.1%} of spending"))

        data.extend([
            (None, None, None, "spacer"),
            ("TOP MERCHANTS", None, None, "section"),
        ])

        # Top merchants
        top_merchants = self._get_top_merchants()
        for merchant, amount, count in top_merchants:
            data.append((merchant, amount, f"{count} transactions"))

        # Account balances preview
        if self.data.has_accounts:
            data.extend([
                (None, None, None, "spacer"),
                ("ACCOUNT BALANCES", None, None, "section"),
            ])
            accounts = self._get_account_balances()
            for name, balance in accounts[:5]:
                data.append((name, balance, ""))

        # Write data
        for item in data:
            if len(item) < 4:
                metric, value, notes = item
                style = None
            else:
                metric, value, notes, style = item

            if style == "section":
                self.formatter.format_section_header(ws, row, metric, 3)
                row += 1
                continue

            if style == "spacer":
                row += 1
                continue

            ws.cell(row=row, column=1, value=metric)

            if value is not None:
                cell = ws.cell(row=row, column=2)
                if style == "total":
                    self.formatter.apply_currency(cell, value, is_total=True)
                elif style == "percent":
                    self.formatter.apply_percent(cell, value)
                else:
                    self.formatter.apply_currency(cell, value)

            ws.cell(row=row, column=3, value=notes)
            row += 1

        self._auto_width(ws)
        return ws

    def _get_date_range(self) -> tuple[str, str]:
        """Get the date range of transactions."""
        df = self.data.transactions
        start = df["date"].min().strftime("%Y-%m-%d")
        end = df["date"].max().strftime("%Y-%m-%d")
        return start, end

    def _calculate_totals(self) -> dict:
        """Calculate period totals."""
        df = self.data.transactions

        income = df[df["amount"] > 0]["amount"].sum()
        expenses = abs(df[df["amount"] < 0]["amount"].sum())
        net = income - expenses
        savings_rate = net / income if income > 0 else 0

        return {
            "income": income,
            "expenses": expenses,
            "net": net,
            "savings_rate": savings_rate,
        }

    def _calculate_net_worth(self) -> tuple[float, float, float]:
        """Calculate net worth from accounts."""
        df = self.data.accounts

        if self.config.exclude_hidden and "is_hidden" in df.columns:
            df = df[df["is_hidden"] != True]

        assets = df[df["current_balance"] >= 0]["current_balance"].sum()
        liabilities = abs(df[df["current_balance"] < 0]["current_balance"].sum())
        net_worth = assets - liabilities

        return assets, liabilities, net_worth

    def _get_top_categories(self) -> list[tuple[str, float, float]]:
        """Get top 5 spending categories."""
        df = self.data.transactions[self.data.transactions["amount"] < 0].copy()

        if df.empty:
            return []

        df["abs_amount"] = df["amount"].abs()
        total = df["abs_amount"].sum()

        category_totals = df.groupby("category_name")["abs_amount"].sum().sort_values(ascending=False).head(5)

        result = []
        for cat, amount in category_totals.items():
            pct = amount / total if total > 0 else 0
            result.append((cat, amount, pct))

        return result

    def _get_top_merchants(self) -> list[tuple[str, float, int]]:
        """Get top 5 merchants by spending."""
        df = self.data.transactions[self.data.transactions["amount"] < 0].copy()

        if df.empty:
            return []

        df["abs_amount"] = df["amount"].abs()

        merchant_stats = df.groupby("merchant_name").agg({"abs_amount": "sum", "id": "count"}).sort_values("abs_amount", ascending=False).head(5)

        result = []
        for merchant, row in merchant_stats.iterrows():
            result.append((merchant, row["abs_amount"], int(row["id"])))

        return result

    def _get_account_balances(self) -> list[tuple[str, float]]:
        """Get account balances sorted by absolute value."""
        df = self.data.accounts

        if self.config.exclude_hidden and "is_hidden" in df.columns:
            df = df[df["is_hidden"] != True]

        result = []
        for _, row in df.iterrows():
            name = row.get("display_name", "Unknown")
            balance = row.get("current_balance", 0) or 0
            result.append((name, balance))

        result.sort(key=lambda x: abs(x[1]), reverse=True)
        return result
