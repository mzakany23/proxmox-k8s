"""Monthly spending pivot sheet."""

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseSheet


class MonthlySheet(BaseSheet):
    """Monthly spending pivot table."""

    sheet_id = "monthly"
    sheet_name = "Monthly Spending"

    def can_generate(self) -> tuple[bool, str]:
        if not self.data.has_transactions:
            return False, "No transaction data available"
        return True, ""

    def generate(self, wb: Workbook) -> Worksheet:
        ws = self._create_sheet(wb)

        # Prepare data - focus on expenses
        df = self.data.transactions[self.data.transactions["amount"] < 0].copy()

        if df.empty:
            ws.cell(row=1, column=1, value="No expense transactions found")
            return ws

        df["abs_amount"] = df["amount"].abs()
        df["month"] = df["date"].dt.to_period("M").astype(str)

        # Create pivot table
        pivot = pd.pivot_table(
            df,
            values="abs_amount",
            index="category_name",
            columns="month",
            aggfunc="sum",
            fill_value=0,
        )

        # Sort by total descending
        pivot["Total"] = pivot.sum(axis=1)
        pivot = pivot.sort_values("Total", ascending=False)

        # Write headers
        row = 1
        ws.cell(row=row, column=1, value="Category").style = "header"

        # Month columns
        months = [col for col in pivot.columns if col != "Total"]
        for i, month in enumerate(months, start=2):
            ws.cell(row=row, column=i, value=month).style = "header"

        # Total column
        total_col = len(months) + 2
        ws.cell(row=row, column=total_col, value="Total").style = "header"

        row += 1

        # Write data rows
        data_start_row = row
        for category, data in pivot.iterrows():
            ws.cell(row=row, column=1, value=category)

            for i, month in enumerate(months, start=2):
                value = data[month]
                cell = ws.cell(row=row, column=i, value=value)
                cell.style = "currency"

            cell = ws.cell(row=row, column=total_col, value=data["Total"])
            cell.style = "currency"

            row += 1

        data_end_row = row - 1

        # Add totals row
        ws.cell(row=row, column=1, value="TOTAL")

        for i, month in enumerate(months, start=2):
            total = pivot[month].sum()
            cell = ws.cell(row=row, column=i, value=total)
            cell.style = "total"

        grand_total = pivot["Total"].sum()
        cell = ws.cell(row=row, column=total_col, value=grand_total)
        cell.style = "total"

        # Also style the label
        ws.cell(row=row, column=1).style = "total"

        # Apply heat map to data cells (excluding totals)
        if data_end_row >= data_start_row:
            self.formatter.apply_heat_map(
                ws,
                start_row=data_start_row,
                end_row=data_end_row,
                start_col=2,
                end_col=total_col - 1,
            )

        # Freeze panes
        self.formatter.freeze_panes(ws, row=2, col=2)

        # Auto-width columns
        self._auto_width(ws)

        return ws
