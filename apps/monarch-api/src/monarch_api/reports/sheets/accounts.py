"""Account summary sheet."""

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .base import BaseSheet


class AccountsSheet(BaseSheet):
    """Account summary with balances and activity."""

    sheet_id = "accounts"
    sheet_name = "Accounts"

    def can_generate(self) -> tuple[bool, str]:
        if not self.data.has_accounts:
            return False, "No account data available"
        return True, ""

    def generate(self, wb: Workbook) -> Worksheet:
        ws = self._create_sheet(wb)

        df = self.data.accounts.copy()

        # Filter hidden accounts if configured
        if self.config.exclude_hidden and "is_hidden" in df.columns:
            df = df[df["is_hidden"] != True]

        if df.empty:
            ws.cell(row=1, column=1, value="No accounts found")
            return ws

        # Calculate inflow/outflow from transactions if available
        account_activity = self._calculate_account_activity()

        # Write header
        row = 1
        headers = ["Account", "Type", "Institution", "Balance", "Inflow", "Outflow", "Net Change"]
        for i, header in enumerate(headers, start=1):
            ws.cell(row=row, column=i, value=header).style = "header"
        row += 1

        # Separate assets and liabilities
        assets = df[df["current_balance"] >= 0].copy()
        liabilities = df[df["current_balance"] < 0].copy()

        totals = {"balance": 0, "inflow": 0, "outflow": 0, "net": 0}

        # Assets section
        if not assets.empty:
            self.formatter.format_section_header(ws, row, "ASSETS", len(headers))
            row += 1

            assets = assets.sort_values("current_balance", ascending=False)
            for _, acc in assets.iterrows():
                row = self._write_account_row(ws, row, acc, account_activity, totals)

        # Liabilities section
        if not liabilities.empty:
            row += 1  # Spacer
            self.formatter.format_section_header(ws, row, "LIABILITIES", len(headers))
            row += 1

            liabilities = liabilities.sort_values("current_balance", ascending=True)
            for _, acc in liabilities.iterrows():
                row = self._write_account_row(ws, row, acc, account_activity, totals)

        # Write totals row
        row += 1
        ws.cell(row=row, column=1, value="NET WORTH")
        ws.cell(row=row, column=1).style = "total"

        cell = ws.cell(row=row, column=4, value=totals["balance"])
        cell.style = "total"

        cell = ws.cell(row=row, column=5, value=totals["inflow"])
        cell.style = "total"

        cell = ws.cell(row=row, column=6, value=totals["outflow"])
        cell.style = "total"

        cell = ws.cell(row=row, column=7, value=totals["net"])
        cell.style = "total"

        # Freeze panes
        self.formatter.freeze_panes(ws, row=2, col=1)

        # Auto-width columns
        self._auto_width(ws)

        return ws

    def _write_account_row(self, ws: Worksheet, row: int, acc, account_activity: dict, totals: dict) -> int:
        """Write a single account row."""
        account_id = acc.get("id")
        name = acc.get("display_name", "Unknown")
        account_type = acc.get("type", "")
        institution = acc.get("institution_name", "")
        balance = acc.get("current_balance", 0) or 0

        # Get activity from transactions
        activity = account_activity.get(account_id, {"inflow": 0, "outflow": 0})
        inflow = activity["inflow"]
        outflow = activity["outflow"]
        net_change = inflow - outflow

        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=account_type)
        ws.cell(row=row, column=3, value=institution)

        cell = ws.cell(row=row, column=4, value=balance)
        self.formatter.apply_currency(cell, balance)

        cell = ws.cell(row=row, column=5, value=inflow)
        cell.style = "currency"

        cell = ws.cell(row=row, column=6, value=outflow)
        cell.style = "currency"

        cell = ws.cell(row=row, column=7, value=net_change)
        self.formatter.apply_currency(cell, net_change)

        totals["balance"] += balance
        totals["inflow"] += inflow
        totals["outflow"] += outflow
        totals["net"] += net_change

        return row + 1

    def _calculate_account_activity(self) -> dict:
        """Calculate inflow/outflow per account from transactions."""
        if not self.data.has_transactions:
            return {}

        df = self.data.transactions

        result = {}

        for account_id in df["account_id"].unique():
            account_txns = df[df["account_id"] == account_id]

            inflow = account_txns[account_txns["amount"] > 0]["amount"].sum()
            outflow = abs(account_txns[account_txns["amount"] < 0]["amount"].sum())

            result[account_id] = {"inflow": inflow, "outflow": outflow}

        return result
