"""Excel formatting utilities."""

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def create_styles(wb: Workbook) -> dict:
    """Create and register named styles for the workbook."""
    styles = {}

    # Header style - dark blue background, white text
    header_style = NamedStyle(name="header")
    header_style.font = Font(bold=True, color="FFFFFF", size=11)
    header_style.fill = PatternFill("solid", fgColor="2F5496")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(
        bottom=Side(style="medium", color="000000"),
    )
    wb.add_named_style(header_style)
    styles["header"] = header_style

    # Section header style - gray background, blue text
    section_style = NamedStyle(name="section")
    section_style.font = Font(bold=True, size=12, color="2F5496")
    section_style.fill = PatternFill("solid", fgColor="D6DCE5")
    section_style.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(section_style)
    styles["section"] = section_style

    # Currency style
    currency_style = NamedStyle(name="currency")
    currency_style.number_format = '"$"#,##0.00'
    currency_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(currency_style)
    styles["currency"] = currency_style

    # Percentage style
    percent_style = NamedStyle(name="percent")
    percent_style.number_format = "0.0%"
    percent_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(percent_style)
    styles["percent"] = percent_style

    # Total row style - green background, double border
    total_style = NamedStyle(name="total")
    total_style.font = Font(bold=True, size=11)
    total_style.fill = PatternFill("solid", fgColor="E2EFDA")
    total_style.number_format = '"$"#,##0.00'
    total_style.border = Border(
        top=Side(style="thin", color="000000"),
        bottom=Side(style="double", color="000000"),
    )
    wb.add_named_style(total_style)
    styles["total"] = total_style

    # Negative currency style (red)
    negative_style = NamedStyle(name="negative")
    negative_style.font = Font(color="C00000")
    negative_style.number_format = '"$"#,##0.00'
    negative_style.alignment = Alignment(horizontal="right")
    wb.add_named_style(negative_style)
    styles["negative"] = negative_style

    # Title style
    title_style = NamedStyle(name="title")
    title_style.font = Font(bold=True, size=16, color="2F5496")
    title_style.alignment = Alignment(horizontal="center")
    wb.add_named_style(title_style)
    styles["title"] = title_style

    return styles


def auto_width(ws: Worksheet, min_width: int = 8, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column = None
        for cell in column_cells:
            try:
                if hasattr(cell, "column_letter"):
                    if column is None:
                        column = cell.column_letter
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        if column:
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column].width = adjusted_width


class ExcelFormatter:
    """Formatting utilities for Excel workbooks."""

    # Color palette matching the analysis project
    COLORS = {
        "header_bg": "2F5496",
        "header_font": "FFFFFF",
        "section_bg": "D6DCE5",
        "section_font": "2F5496",
        "total_bg": "E2EFDA",
        "negative": "C00000",
    }

    @classmethod
    def format_header_row(cls, ws: Worksheet, row: int, start_col: int = 1, end_col: int | None = None) -> None:
        """Apply header formatting to a row."""
        if end_col is None:
            end_col = ws.max_column

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.style = "header"

    @classmethod
    def format_section_header(cls, ws: Worksheet, row: int, text: str, end_col: int) -> None:
        """Format a section header with merged cells."""
        cell = ws.cell(row=row, column=1, value=text)
        cell.style = "section"
        if end_col > 1:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)

    @classmethod
    def format_total_row(cls, ws: Worksheet, row: int, start_col: int = 1, end_col: int | None = None) -> None:
        """Apply total row formatting."""
        if end_col is None:
            end_col = ws.max_column

        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.style = "total"

    @classmethod
    def apply_currency(cls, cell, value: float, is_total: bool = False) -> None:
        """Apply currency formatting to a cell."""
        cell.value = value
        if is_total:
            cell.style = "total"
        elif value < 0:
            cell.style = "negative"
        else:
            cell.style = "currency"

    @classmethod
    def apply_percent(cls, cell, value: float) -> None:
        """Apply percentage formatting to a cell."""
        cell.value = value
        cell.style = "percent"

    @classmethod
    def freeze_panes(cls, ws: Worksheet, row: int = 2, col: int = 1) -> None:
        """Freeze rows and columns."""
        ws.freeze_panes = ws.cell(row=row, column=col)

    @classmethod
    def apply_heat_map(
        cls, ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int
    ) -> None:
        """Apply heat map coloring based on cell values."""
        values = []
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    values.append(cell.value)

        if not values:
            return

        min_val = min(values)
        max_val = max(values)
        val_range = max_val - min_val if max_val != min_val else 1

        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    intensity = (cell.value - min_val) / val_range
                    # Interpolate from white to light red
                    r = 255
                    g = int(255 - (255 - 200) * intensity)
                    b = int(255 - (255 - 200) * intensity)
                    color = f"{r:02X}{g:02X}{b:02X}"
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
