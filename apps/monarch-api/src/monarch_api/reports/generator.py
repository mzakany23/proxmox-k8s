"""Main Excel report generator."""

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .config import ReportConfig
from .data_loader import DataLoader, LoadedData
from .formatters import ExcelFormatter, auto_width, create_styles
from .sheets import (
    AccountsSheet,
    BaseSheet,
    CategoriesSheet,
    MonthlySheet,
    SummarySheet,
    TransactionsSheet,
)


class ExcelReportGenerator:
    """Generate Excel financial reports from Monarch Money CSV exports."""

    SHEET_CLASSES: dict[str, type[BaseSheet]] = {
        "summary": SummarySheet,
        "monthly": MonthlySheet,
        "categories": CategoriesSheet,
        "accounts": AccountsSheet,
        "transactions": TransactionsSheet,
    }

    def __init__(self, config: ReportConfig):
        self.config = config
        self.data: LoadedData | None = None
        self.skipped_sheets: list[tuple[str, str]] = []
        self.styles: dict = {}

    def generate(self) -> Path:
        """Generate the Excel report.

        Returns:
            Path to the generated Excel file
        """
        # Load data
        loader = DataLoader(self.config)
        self.data = loader.load()

        # Create workbook and register styles
        wb = Workbook()
        self.styles = create_styles(wb)

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Generate requested sheets
        generated_sheets = []
        for sheet_id in self.config.sheets:
            if sheet_id not in self.SHEET_CLASSES:
                print(f"Warning: Unknown sheet type '{sheet_id}', skipping")
                continue

            sheet_class = self.SHEET_CLASSES[sheet_id]
            sheet = sheet_class(self.config, self.data, self.styles)

            can_gen, reason = sheet.can_generate()
            if can_gen:
                sheet.generate(wb)
                generated_sheets.append(sheet_id)
                print(f"  Generated: {sheet.sheet_name}")
            else:
                self.skipped_sheets.append((sheet.sheet_name, reason))
                print(f"  Skipped: {sheet.sheet_name} ({reason})")

        # Add metadata sheet if any sheets were skipped
        if self.skipped_sheets or self.data.missing_files:
            self._add_metadata_sheet(wb)

        # Determine output path
        output_path = self._get_output_path()

        # Save workbook
        wb.save(output_path)

        return output_path

    def _add_metadata_sheet(self, wb: Workbook) -> Worksheet:
        """Add a metadata sheet with report info and skipped sheets."""
        ws = wb.create_sheet(title="Report Info")

        row = 1

        # Title
        ws.merge_cells("A1:B1")
        ws.cell(row=row, column=1, value="Report Metadata")
        ws.cell(row=row, column=1).style = "title"
        row += 2

        # Generation info
        ws.cell(row=row, column=1, value="Generated:")
        ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row += 1

        ws.cell(row=row, column=1, value="Export Directory:")
        ws.cell(row=row, column=2, value=str(self.config.export_dir))
        row += 2

        # Data files section
        ExcelFormatter.format_section_header(ws, row, "AVAILABLE DATA FILES", 2)
        row += 1
        for filename in self.data.available_files:
            ws.cell(row=row, column=1, value=filename)
            row += 1
        row += 1

        # Missing files
        if self.data.missing_files:
            ExcelFormatter.format_section_header(ws, row, "MISSING DATA FILES", 2)
            row += 1
            for filename in self.data.missing_files:
                ws.cell(row=row, column=1, value=filename)
                row += 1
            row += 1

        # Skipped sheets
        if self.skipped_sheets:
            ExcelFormatter.format_section_header(ws, row, "SKIPPED SHEETS", 2)
            row += 1

            ws.cell(row=row, column=1, value="Sheet")
            ws.cell(row=row, column=2, value="Reason")
            ExcelFormatter.format_header_row(ws, row, 1, 2)
            row += 1

            for sheet_name, reason in self.skipped_sheets:
                ws.cell(row=row, column=1, value=sheet_name)
                ws.cell(row=row, column=2, value=reason)
                row += 1

        auto_width(ws)

        return ws

    def _get_output_path(self) -> Path:
        """Determine the output file path."""
        if self.config.output_path:
            return self.config.output_path

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"monarch_report_{timestamp}.xlsx"

        return self.config.export_dir / filename
